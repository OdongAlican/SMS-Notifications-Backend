from celery import shared_task
from pride_notify_notice.utils import handle_ATM_expiry, handle_Escrow_notifications, handle_loans_due, handle_birthdays, handle_URA_reports, handle_group_loans, update_ATM_expiry, update_group_loans
import urllib3
from datetime import datetime
import json
from .models import ATMExpirySMSLog, GroupLoanSMSLog, SMSLog, BirthdaySMSLog, GroupSMSLog
from dateutil.parser import parse
import os
from dotenv import load_dotenv
from django.conf import settings
from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.core.mail import EmailMessage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
load_dotenv()

@shared_task(bind=True, max_retries=5, default_retry_delay=300)
def retrieve_data(self):
    try:
        loan_data = handle_loans_due()
        # print(loan_data)
        person_list = loan_data.get("Person", [])

        if not person_list:
            raise ValueError("Empty 'Person' list received.")

        # updated_loan_list = update_List(person_list)
        response_data = []

        for loan in person_list:
            response = send_sms_to_api(loan)
            if response:
                response_data.append(response)

        return response_data

    except (ValueError) as e:
        print(f"Data error: {e}")
        raise self.retry(exc=e)

    except Exception as exc:
        print(f"Unexpected error occurred: {exc}")
        raise self.retry(exc=exc)



@shared_task(bind=True, max_retries=5, default_retry_delay=300)
def retrieve_birthday_data(self):
    try:
        birthday_data = handle_birthdays()
        # print(birthday_data)
        person_list = birthday_data.get("Person", [])

        if not person_list:
            raise ValueError("Empty 'Person' list received.")

        # updated_birthday_list = update_List_birthdays(person_list)
        response_data = []

        for birthday in person_list:
            response = send_sms_to_api(birthday)
            if response:
                response_data.append(response)

        return response_data

    except (ValueError) as e:
        print(f"Data error: {e}")
        raise self.retry(exc=e)

    except Exception as exc:
        print(f"Unexpected error occurred: {exc}")
        raise self.retry(exc=exc)


@shared_task(bind=True, max_retries=5, default_retry_delay=300)
def retrieve_atm_expiry_notifications(self):
    try:
        atm_expiry_data = handle_ATM_expiry()
        # print(atm_expiry_data)
        person_list = atm_expiry_data.get("Person", [])

        if not person_list:
            raise ValueError("Empty 'Person' list received.")

        updated_atm_expiry_list = update_ATM_expiry(person_list)

        response_data = []
        for atm_expiry in updated_atm_expiry_list:
            response = send_sms_to_api(atm_expiry)
            if response:
                response_data.append(response)

        return response_data

    except (ValueError) as e:
        print(f"Data error: {e}")
        raise self.retry(exc=e)

    except Exception as exc:
        print(f"Unexpected error occurred: {exc}")
        raise self.retry(exc=exc)
    
    
@shared_task(bind=True, max_retries=5, default_retry_delay=300)
def retrieve_escrow_notifications(self):
    try:
        escrow_data = handle_Escrow_notifications()
        print(escrow_data)

        # Normalize input to a list of transactions
        if isinstance(escrow_data, list):
            notifications = escrow_data
        elif isinstance(escrow_data, dict):
            notifications = (
                escrow_data.get("statement")
                or escrow_data.get("Report")
                or escrow_data.get("data")
                or []
            )
        else:
            notifications = []

        if not notifications:
            raise ValueError("Empty 'Notifications' list received.")

        # Build Excel workbook borrowing layout from provided PDF
        wb = Workbook()
        ws = wb.active
        ws.title = "MTN Escrow Statement"

        # Extract header-level details from first record where available
        first = next((n for n in notifications if isinstance(n, dict)), {})
        acct_name = (first.get('ACCT_NM') or '').strip()
        address = (first.get('ADDR_LINE_1') or '').strip()
        branch_name = (first.get('BU_NM') or '').strip()
        account_no = (first.get('ACT_NO') or '').strip()
        product = (first.get('PROD_DESC') or '').strip()
        currency = (first.get('CRNCY_NM') or first.get('CRNCY_CD_ISO') or '').strip()
        bank_name = (first.get('BANK_NAME') or 'Pride Bank').strip()

        # Determine date range
        def safe_parse_date(dt):
            try:
                if not dt:
                    return None
                return parse(str(dt))
            except Exception:
                return None

        dates = [safe_parse_date(n.get('TRAN_DT')) for n in notifications if isinstance(n, dict)]
        dates = [d for d in dates if d is not None]
        from_date = min(dates).strftime('%d/%m/%Y') if dates else ''
        to_date = max(dates).strftime('%d/%m/%Y') if dates else ''
        printed_on = datetime.now().strftime('%d/%m/%Y')

        # Opening/Closing and totals (fallbacks if missing)
        def to_float(val):
            try:
                if val in (None, ''):
                    return 0.0
                return float(str(val).replace(',', ''))
            except Exception:
                return 0.0

        opening_balance = to_float(first.get('OPENIING_BAL')) or to_float(first.get('OPENING_BAL'))
        # closing_balance = to_float(first.get('CLOSING_BAL')) or (to_float(notifications[-1].get('STMNT_BAL')) if notifications else 0.0)

        total_debits = sum(to_float(n.get('DEBIT_AMT')) for n in notifications)
        total_credits = sum(to_float(n.get('CREDIT_AMT')) for n in notifications)
        count_debits = sum(1 for n in notifications if to_float(n.get('DEBIT_AMT')) > 0)
        count_credits = sum(1 for n in notifications if to_float(n.get('CREDIT_AMT')) > 0)

        # Title
        ws.append(["MTN Escrow Transaction Statement"])
        ws.merge_cells('A1:N1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Modern, two-column header block (clean professional styling)
        # Left column fields
        left_rows = [
            ("Acct Name", acct_name),
            ("Address", address),
            ("Branch Name", branch_name),
            ("Account No", account_no),
            ("Product", product),
        ]
        # Right column fields
        right_rows = [
            ("Currency", currency),
            ("From Date", from_date),
            ("To Date", to_date),
            ("Bank", bank_name),
            ("Printed On", printed_on),
        ]

        # Column layout: A(Label) B-C(Value) D(Spacer) E(Label) F-H(Value)
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 6
        ws.column_dimensions['D'].width = 4
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 6
        ws.column_dimensions['H'].width = 4

        label_fill = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
        value_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        header_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
        )

        start_row = 3
        # Render left column
        for i, (label, value) in enumerate(left_rows):
            r = start_row + i
            # Label cell
            lc = ws.cell(row=r, column=1, value=f"{label}:")
            lc.font = Font(bold=True)
            lc.alignment = Alignment(horizontal='left', vertical='center')
            lc.fill = label_fill
            lc.border = header_border
            # Value cells (merge B:C)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            vc = ws.cell(row=r, column=2, value=value)
            vc.alignment = Alignment(horizontal='left', vertical='center')
            vc.fill = value_fill
            vc.border = header_border

        # Render right column
        for i, (label, value) in enumerate(right_rows):
            r = start_row + i
            rc = ws.cell(row=r, column=5, value=f"{label}:")
            rc.font = Font(bold=True)
            rc.alignment = Alignment(horizontal='left', vertical='center')
            rc.fill = label_fill
            rc.border = header_border
            ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
            rv = ws.cell(row=r, column=6, value=value)
            rv.alignment = Alignment(horizontal='left', vertical='center')
            rv.fill = value_fill
            rv.border = header_border

        # Add subtle spacing below header block
        ws.append([""])

        # Opening balance line
        ws.append([""])
        ws.append([f"Opening balance : {opening_balance:,.2f}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
        ws[ f'A{ws.max_row}' ].font = Font(bold=True)

        # Table headers
        ws.append([""])
        headers = [
            "Transaction Date", "Value Date", "Bank Reference", "MTN Reference", "MSISDN",
            "Transaction Description", "Dr / Cr", "Debit", "Credit", "Balance",
            "CBS Status", "Prefunding", "Posted By", "Branch"
        ]
        ws.append(headers)
        header_row_idx = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        amount_cols = {8, 9, 10}  # Debit, Credit, Balance

        # Ensure chronological order
        def sort_key(n):
            dt = safe_parse_date(n.get('TRAN_DT')) if isinstance(n, dict) else None
            return dt or datetime.min
        notifications_sorted = sorted(notifications, key=sort_key)

        # Initialize running balance from opening balance
        running_balance = opening_balance

        for n in notifications_sorted:
            if not isinstance(n, dict):
                continue
            tran_dt = safe_parse_date(n.get('TRAN_DT'))
            value_dt = safe_parse_date(n.get('VALUE_DT'))
            tran_dt_s = tran_dt.strftime('%d/%m/%Y') if tran_dt else ''
            value_dt_s = value_dt.strftime('%d/%m/%Y') if value_dt else ''

            tran_desc = n.get('TRAN_DESC') or ''
            reference = n.get('TRAN_REF_TXT') or ''
            bank_ref = n.get('SETTLEMENT_BANK_REF') or ''
            drcr = (n.get('DR_CR_IND') or ('DR' if to_float(n.get('DEBIT_AMT')) > 0 else 'CR' if to_float(n.get('CREDIT_AMT')) > 0 else '')).strip()
            debit = to_float(n.get('DEBIT_AMT'))
            credit = to_float(n.get('CREDIT_AMT'))

            # Calculate running balance: opening + credits - debits
            running_balance = running_balance + credit - debit
            balance = running_balance
            cbs_status = n.get('CBS_Status') or ''
            prefunding = (n.get('PREFUNDING_BRANCH') or '').strip()
            posted_by = (n.get('POSTED_BY') or n.get('USER_NAME') or '').strip()
            branch = (n.get('BU_NM') or '').strip()
            msisdn = (n.get('CONTACT') or '').strip()

            row = [
                tran_dt_s,
                value_dt_s,
                str(reference),
                str(bank_ref),
                msisdn,
                tran_desc,
                drcr,
                debit,
                credit,
                balance,
                cbs_status,
                prefunding,
                posted_by,
                branch,
            ]
            ws.append(row)

        # Format numeric columns
        for r in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for idx, cell in enumerate(r, start=1):
                if idx in amount_cols:
                    cell.number_format = '#,##0.00'

        # Auto-fit columns
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        # Borders and zebra stripes
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=header_row_idx, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")

        # Summary footer similar to PDF
        ws.append([""])
        ws.append([f"Debit(s) - {count_debits}  Credit(s) - {count_credits}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
        ws[ f'A{ws.max_row}' ].font = Font(bold=True)

        # Totals line: place values under Debit/Credit/Balance columns
        totals_row_idx = ws.max_row + 1
        ws.cell(row=totals_row_idx, column=1, value="Total :- ").font = Font(bold=True)
        ws.cell(row=totals_row_idx, column=8, value=total_debits).number_format = '#,##0.00'
        ws.cell(row=totals_row_idx, column=9, value=total_credits).number_format = '#,##0.00'
        ws.cell(row=totals_row_idx, column=10, value=(running_balance if notifications_sorted else opening_balance)).number_format = '#,##0.00'

        # Closing balance line
        ws.append([""])
        ws.append([f"Closing { (running_balance if notifications_sorted else opening_balance):,.2f}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
        ws[ f'A{ws.max_row}' ].font = Font(bold=True)

        # Footer: printed by / verified by (dummy where missing)
        # printed_by = next((n.get('POSTED_BY') for n in notifications_sorted if (n.get('POSTED_BY') or '').strip()), None) or (first.get('USER_NAME') or 'SYSTEM SYSTEM')
        ws.append([""])
        # ws.append([f"Printed By : {printed_by}"])
        ws.append([f"Printed By : CUSTOMER ENGAGEMENT SYSTEM"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
        ws.append([f"Print Date: {datetime.now().strftime('%d-%b-%Y')} "])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
        ws.append(["Verified By: CUSTOMER ENGAGEMENT SYSTEM"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)

        # Save file
        excel_filename = f"mtn_escrow_statement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(excel_filename)

        # Send the Excel file via email (same helper used by URA report)
        send_csv_report_email(
            recipient_email=getattr(settings, 'ESCROW_REPORT_EMAILS', []),
            subject=f"Daily MTN Escrow Statement - {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}",
            message="Dear Valued Partner, \nPlease find the attached MTN Escrow statement.",
            csv_file_path=excel_filename
        )

        print(f"Escrow statement saved to {excel_filename}")
        return [{
            'filename': excel_filename,
            'content': f"Escrow statement generated and saved to {excel_filename}",
            'totals': {
                'debits': total_debits,
                'credits': total_credits,
                'count_debits': count_debits,
                'count_credits': count_credits,
            }
        }]

    except (ValueError) as e:
        print(f"Data error: {e}")
        raise self.retry(exc=e)

    except Exception as exc:
        print(f"Unexpected error occurred: {exc}")
        raise self.retry(exc=exc)

@shared_task(bind=True, max_retries=5, default_retry_delay=300)
def retrieve_ura_report(self):
    try:
        ura_report_data = handle_URA_reports()
        print(ura_report_data)

        report_list = ura_report_data.get("Report", [])
        if not report_list:
            raise ValueError("Empty 'Report' list received.")

        wb = Workbook()
        ws = wb.active
        ws.title = "URA Report"
        
        # Get the initial ledger balance from the first transaction for opening balance
        opening_balance = float(report_list[0].get('OPENING_BALANCE', 0)) if report_list else 0
        
        # Extract GL Account Number from first item
        gl_account_no = report_list[0].get('GL_ACCT_NO', 'N/A') if report_list else 'N/A'
        
        # Get start and end dates from first and last items in report
        try:
            start_date = datetime.fromisoformat(report_list[0]['TRAN_DT'].replace('Z', '+00:00')).strftime('%d/%m/%Y') if report_list else 'N/A'
            end_date = datetime.fromisoformat(report_list[-1]['TRAN_DT'].replace('Z', '+00:00')).strftime('%d/%m/%Y') if report_list else 'N/A'
        except (ValueError, KeyError, TypeError):
            start_date = 'N/A'
            end_date = 'N/A'
        
        # Add report title and date
        report_date = datetime.now().strftime('%d/%m/%Y')
        
        # Insert title row (Row 1)
        ws.append(["URA Transaction Report"])
        ws.merge_cells(f'A1:M1')  # Merge cells across all columns
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Insert date row (Row 2)
        ws.append([f"Generated on: {report_date}"])
        ws.merge_cells(f'A2:M2')  # Merge cells across all columns
        date_cell = ws['A2']
        date_cell.font = Font(italic=True)
        date_cell.alignment = Alignment(horizontal='center')
        
        # Insert GL Account No row (Row 3)
        ws.append([f"GL Account No: {gl_account_no}"])
        ws.merge_cells(f'A3:M3')  # Merge cells across all columns
        gl_cell = ws['A3']
        gl_cell.font = Font(bold=True)
        gl_cell.alignment = Alignment(horizontal='left')
        
        # Insert Account Description row (Row 4)
        ws.append(["Uganda Revenue Authority ePayments -Cash"])
        ws.merge_cells(f'A4:M4')  # Merge cells across all columns
        desc_cell = ws['A4']
        desc_cell.alignment = Alignment(horizontal='left')
        
        # Insert Date Range row (Row 5)
        ws.append([f"Period: {start_date} to {end_date}"])
        ws.merge_cells(f'A5:M5')  # Merge cells across all columns
        range_cell = ws['A5']
        range_cell.font = Font(bold=True)
        range_cell.alignment = Alignment(horizontal='left')
        
        # Insert empty row (Row 6)
        ws.append([""])
        
        # Insert opening balance section (Row 7)
        opening_balance_formatted = "{:,.2f}".format(opening_balance)
        ws.append(["Opening Balance:", "", "", "", "", "", "", "", "", "", "", opening_balance_formatted, ""])
        ws.merge_cells(f'A7:K7')  # Merge cells A-K
        opening_balance_label = ws['A7']
        opening_balance_label.font = Font(bold=True)
        opening_balance_label.alignment = Alignment(horizontal='right')
        
        # Format the opening balance cell
        opening_balance_cell = ws['L7']  # Column L, row 7
        opening_balance_cell.font = Font(bold=True)
        opening_balance_cell.number_format = '#,##0.00'
        
        # Insert empty row (Row 8)
        ws.append([""])
        
        # Add headers (Row 9)
        headers = [
            "S/N", "Post Date", "Effective Date", "Created By", "Transaction Description",
            "Reference", "Contra Account", "Origin Branch", "Debit", "Credit", "PRN", "Balance", "Payment Mode"
        ]
        ws.append(headers)

        # Bold headers
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=9, column=col_num).font = Font(bold=True)  # Now header is on row 9
            ws.cell(row=9, column=col_num).alignment = Alignment(horizontal='center')

        # Initialize running balance with opening balance
        running_balance = opening_balance

        # Add transactions starting from row 10
        start_row = 10
        for idx, report in enumerate(report_list, 1):
            try:
                post_date = datetime.fromisoformat(report['TRAN_DT'].replace('Z', '+00:00')).strftime('%d/%m/%Y')
            except:
                post_date = ""

            try:
                effective_date = datetime.fromisoformat(report['EFFECTIVE_DT'].replace('Z', '+00:00')).strftime('%d/%m/%Y')
            except:
                effective_date = ""

            created_by = report.get('USER_NAME', '')
            if created_by:
                created_by = ' '.join(word.capitalize() for word in created_by.lower().split())

            transaction_desc = report.get('TRAN_DESC', '')
            prn_value = report.get('PRN', '')
            tin_value = report.get('TIN', '')

            if prn_value and f"PRN: {prn_value}" not in transaction_desc:
                transaction_desc += f", PRN: {prn_value}"
            if tin_value and f"TIN: {tin_value}" not in transaction_desc:
                transaction_desc += f", TIN: {tin_value}"

            debit_amt = float(report.get('DEBIT_AMT') or 0)
            credit_amt = float(report.get('CREDIT_AMT') or 0)
            
            # Calculate running balance
            running_balance = running_balance + credit_amt - debit_amt
            calculated_ledger_bal = running_balance

            row = [
                idx,
                post_date,
                effective_date,
                created_by,
                transaction_desc,
                str(report.get('TRAN_REF_TXT', '')),    # Keep as string
                str(report.get('CONTRA_ACCT_NO', '')),  # Keep as string
                str(report.get('USER_BU', '')),
                debit_amt,
                credit_amt,
                str(prn_value),                         # Force as string
                calculated_ledger_bal,
                str(report.get('PAYMENT_TYPE', '')),
            ]

            ws.append(row)

        # Format numeric columns with comma and 2 decimals - adjust row index for headers now at row 9
        money_columns = ['I', 'J', 'L']  # Debit, Credit, Balance
        for col in money_columns:
            for cell in ws[col][start_row-1:]:  # Start after headers
                cell.number_format = '#,##0.00'

        # Format text columns to prevent scientific notation - adjust row index
        text_columns = ['F', 'G', 'K']  # Reference, Contra Account, PRN
        for col in text_columns:
            for cell in ws[col][start_row-1:]:  # Start after headers
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left')

        # Auto-adjust column widths
        for col_cells in ws.columns:
            max_length = 0
            column = col_cells[0].column
            col_letter = get_column_letter(column)
            for cell in col_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Add borders to the entire table
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
                             
        # Apply the border to all cells with content
        for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=13):
            for cell in row:
                cell.border = thin_border
                
        # Apply alternating row colors for better readability
        for row_idx in range(10, ws.max_row + 1):  # Start from row 10
            if row_idx % 2 == 0:  # even rows
                for col_idx in range(1, 14):  # columns A-M
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="F2F2F2", 
                                                                           end_color="F2F2F2",
                                                                           fill_type="solid")

        # Save Excel file
        excel_filename = f"ura_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(excel_filename)

        # Send the Excel file via email
        send_csv_report_email(
            recipient_email=settings.URA_REPORT_EMAILS,
            subject="URA Report Excel",
            message="Please find attached the latest URA report.",
            csv_file_path=excel_filename
        )

        print(f"URA report saved to {excel_filename}")
        return [{
            'filename': excel_filename,
            'content': f"URA report generated and saved to {excel_filename}"
        }]

    except (ValueError) as e:
        print(f"Data error: {e}")
        raise self.retry(exc=e)

    except Exception as exc:
        print(f"Unexpected error occurred: {exc}")
        raise self.retry(exc=exc)


def send_csv_report_email(recipient_email, subject, message, csv_file_path):
    try:
        recipients = recipient_email if isinstance(recipient_email, list) else [recipient_email]

        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.SENDER_EMAIL,
            to=recipients,
        )

        # Attach CSV file
        if os.path.exists(csv_file_path):
            with open(csv_file_path, 'rb') as f:
                email.attach(os.path.basename(csv_file_path), f.read(), 'text/csv')
        else:
            print("CSV file not found at:", csv_file_path)
            return False

        email.send()
        print(f"Email sent to {recipient_email} with CSV attachment.")
        return True
    except Exception as e:
        print("Error sending CSV report email:", e)
        return False



@shared_task(bind=True, max_retries=5, default_retry_delay=300)
def retrieve_group_loans(self):
    try:
        group_loans_data = handle_group_loans()
        # print("Group Loans Datat",group_loans_data)
        person_list = group_loans_data.get("Report", [])
        # print("Filtered Persons List",person_list)

        if not person_list:
            raise ValueError("Empty 'Person' list received.")

        # updated_birthday_list = update_group_loans(person_list)
        # print(updated_birthday_list)
        response_data = []

        for group_loan in person_list:
            response = send_sms_to_api(group_loan)
            if response:
                response_data.append(response)

        return response_data

    except (ValueError) as e:
        print(f"Data error: {e}")
        raise self.retry(exc=e)

    except Exception as exc:
        print(f"Unexpected error occurred: {exc}")
        raise self.retry(exc=exc)


@shared_task(bind=True, max_retries=2, default_retry_delay=60)
def send_sms_to_api(self, message_detail):
    resp = ""
    http = urllib3.PoolManager(cert_reqs='CERT_NONE', assert_hostname=False)
    response_data = {}
    
    try:
        # Set up base fields
        acct_nm = ""
        tel_number = ""
        message = ""
        log_model = None
        due_dt_for_db = None
        amt_due = None
        date_of_birth = None
        group_cust_no = None

        # Handle custom messages (New condition for insurance messages)
        if 'CUSTOM_MESSAGE' in message_detail:
            acct_nm = message_detail.get('CUST_NM')
            tel_number = message_detail.get('TEL_NUMBER')
            # Use the custom message directly
            message = message_detail.get('CUSTOM_MESSAGE')
            log_model = GroupSMSLog  # Use standard SMS log for custom messages

        elif 'AMT_DUE' in message_detail:
            # Loan due message handling (existing code)
            acct_nm = message_detail.get('CUST_NM')
            tel_number = message_detail.get('TEL_NUMBER')
            due_dt_serial = message_detail.get('DUE_DT')
            amt_due = float(message_detail.get('AMT_DUE', 0))

            due_dt_obj = parse(due_dt_serial) if isinstance(due_dt_serial, str) else due_dt_serial
            due_dt_for_db = due_dt_obj.date()
            due_dt_for_sms = due_dt_obj.strftime('%d-%m-%Y')

            formatted_amt_due = "{:,}".format(round(amt_due))
            message = (
                f"Dear {acct_nm}, your loan instalment is due on {due_dt_for_sms}. "
                "Thank you for banking with us. Toll Free: 0800333999"
            )
            log_model = SMSLog

        elif 'BIRTH_DT' in message_detail:
            # Birthday message
            acct_nm = message_detail.get('FIRST_NM')
            tel_number = message_detail.get('TEL_NUMBER')
            date_of_birth_raw = message_detail.get('BIRTH_DT')
            client_type = message_detail.get('CLIENT_TYPE', 'CUSTOMER')

            try:
                date_of_birth = parse(date_of_birth_raw).date()
            except Exception as e:
                print(f"Failed to parse BIRTH_DT: {e}")
                date_of_birth = None

            message = (
                f"Dear {acct_nm}, Pride Wishes you a Happy Birthday. We value our relationship with you. "
                "Thank you for choosing Pride. Toll Free: 0800333999"
            )
            log_model = BirthdaySMSLog

        
        elif 'GROUP_CUST_NO' in message_detail:
            # Group loan message - updated format
            full_name = message_detail.get('MEMBER_NM', '')
            
            name_parts = full_name.split()
            if len(name_parts) > 1:
                display_name = name_parts[-1]
            else:
                display_name = full_name
        
            acct_nm = full_name
            tel_number = message_detail.get('PHONE', '')
            group_cust_no = message_detail.get('GROUP_CUST_NO', '')
            
            try:
                loan_amount = float(message_detail.get('LOAN_AMOUNT_PAID', '0'))
                comp_amount = float(message_detail.get('COMP_AMOUNT_PAID', '0'))
                vol_amount = float(message_detail.get('VOL_AMOUNT_PAID', '0'))
                
                total_amount = loan_amount + comp_amount + vol_amount
                
                formatted_total = "{:,.0f}".format(total_amount)
            except (ValueError, TypeError) as e:
                print(f"Error calculating total amount: {e}")
                formatted_total = "0"
            
            try:
                create_dt = datetime.fromisoformat(message_detail.get('CREATE_DT', '').replace('Z', '+00:00'))
                formatted_date = create_dt.strftime('%d-%m-%Y')
            except (ValueError, TypeError, AttributeError) as e:
                print(f"Error formatting date: {e}")
                formatted_date = datetime.now().strftime('%d-%m-%Y')
            
            message = (
                f"Dear {display_name}, your CEC/MEC collection of Shs {formatted_total} has been received on {formatted_date}. "
                "For Help Call 0800333999. Never share your ATM/Mobile PIN."
            )
            print("Generated group loan message:", message)
            log_model = GroupLoanSMSLog

        elif "CARD_TITLE" in message_detail:
            # ATM card expiry message
            pan = message_detail.get('PAN_MASKED')
            tel_number = message_detail.get('MOBILE_CONTACT')
            card_title = message_detail.get('CARD_TITLE').split()[0]
            card_title = card_title.replace(',', '')  # Remove comma if present
    
            try:
                date_of_birth = parse(date_of_birth_raw).date()
            except Exception as e:
                print(f"Failed to parse BIRTH_DT: {e}")
                date_of_birth = None

            message = (
                f"Dear {card_title}, your ATM card ending with **{pan[-4:]} will expire at the end of this month. "
                "Please visit your nearest branch to renew. Never share your PIN."
            )
            log_model = ATMExpirySMSLog

        else:
            raise ValueError("Invalid message_detail format: no recognized fields found.")

        # Send SMS (existing code)
        sender_name = os.getenv("MOONLIGHT_SENDER_NAME", "default_sender")
        password = os.getenv("MOONLIGHT_SENDER_PASSWORD", "default_password")
        address = os.getenv("MOONLIGHT_SENDER_ADDRESS", "http://example.com/api")

        resp = http.request(
            'GET',
            f"{address}?sender_name={sender_name}&password={password}&recipient_addr={tel_number}&message={message}"
        )

        # Attempt to parse response
        try:
            api_response = json.loads(resp.data.decode('utf-8'))
        except json.decoder.JSONDecodeError:
            api_response = {"raw_response": resp.data.decode('utf-8')}


        response_data = {
            'account_name': acct_nm,
            'phone_number': tel_number,
            'message': message,
            'due_date': due_dt_for_db,
            'amount_due': amt_due,
            'status': api_response,
            'response_data': api_response
        }

        # Save to appropriate model
        if log_model == SMSLog:
            log_model.objects.create(
                account_name=acct_nm,
                phone_number=tel_number,
                message=message,
                due_date=due_dt_for_db,
                amount_due=amt_due or 0,
                status=api_response,
                response_data=api_response
            )
        # Save to appropriate model
        elif log_model == GroupSMSLog:
            log_model.objects.create(
                account_name=acct_nm,
                phone_number=tel_number,
                message=message,
                status=api_response,
                response_data=api_response
            )
        elif log_model == BirthdaySMSLog:
            log_model.objects.create(
                acct_nm=acct_nm,
                client_type=client_type,
                message=message,
                date_of_birth=date_of_birth,
                contact=tel_number,
                status=api_response,
                response_data=api_response
            )
        elif log_model == GroupLoanSMSLog:
            # Create new group loan SMS log entry
            log_model.objects.create(
                acct_nm=acct_nm,
                group_cust_no=group_cust_no,
                message=message,
                contact=tel_number,
                status=api_response,
                response_data=api_response
            )
        elif log_model == ATMExpirySMSLog:
            log_model.objects.create(
                acct_nm=card_title,
                message=message,
                contact=tel_number,
                status=api_response,
                response_data=api_response
            )

        return response_data

    except Exception as e:
        error_msg = str(e)
        fallback_response = resp.data.decode('utf-8') if resp else "No response received"
        print(f"Error sending SMS: {error_msg}")

        # Log even failed attempts
        if 'SMSLog' in str(type(log_model)):
            log_model.objects.create(
                account_name=acct_nm,
                phone_number=tel_number,
                message=message,
                due_date=due_dt_for_db,
                amount_due=amt_due,
                status=fallback_response,
                response_data={"error": error_msg}
            )
        elif 'BirthdaySMSLog' in str(type(log_model)):
            log_model.objects.create(
                acct_nm=acct_nm,
                client_type=client_type if 'client_type' in locals() else 'Birthday',
                message=message,
                date_of_birth=date_of_birth,
                contact=tel_number,
                status=fallback_response,
                response_data={"error": error_msg}
            )
        elif 'GroupLoanSMSLog' in str(type(log_model)):
            # Error handling for group loan SMS
            log_model.objects.create(
                acct_nm=acct_nm,
                group_cust_no=group_cust_no if group_cust_no else "",
                message=message,
                contact=tel_number,
                status=fallback_response,
                response_data={"error": error_msg}
            )

        return {
            'account_name': acct_nm,
            'phone_number': tel_number,
            'message': message,
            'due_date': due_dt_for_db,
            'amount_due': amt_due,
            'status': fallback_response,
            'response_data': {"error": error_msg}
        }